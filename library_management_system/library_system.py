import os
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from tkinter import filedialog
import pickle
import time
from windowui import *
import requests
import json
from io import BytesIO
import pandas as pd
from datetime import datetime, timedelta
import threading
import re
from difflib import SequenceMatcher
import qrcode
from PIL import Image, ImageTk
# 使用opencv来进行扫码识别
import cv2

import serial
import serial.tools.list_ports
from adafruit_pn532.uart import PN532_UART

import customtkinter as ctk

# Add functools.wraps for decorator usage
from functools import wraps

# 设置 CustomTkinter 主题
ctk.set_appearance_mode("System")       # 系统自动暗黑/亮色
ctk.set_default_color_theme("blue")     # 主题色


# 会话持久化文件路径
session_file = os.path.join(os.path.dirname(__file__), 'session.pkl')

# 持久化借阅历史文件路径
history_file = os.path.join(os.path.dirname(__file__), 'borrow_history.csv')

# 记录操作日志
from datetime import datetime
def log_action(entry: str):
    lp = os.path.join(os.path.dirname(__file__), 'log.txt')
    with open(lp, 'a', encoding='utf-8') as f:
        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {entry}\n")

listbook = []
power = False
# 管理员会话开始时间
session_start = None


# 判断管理员是否登录
def log(func):
    global power
    @wraps(func)
    def wrapper(*args, **kw):
        from datetime import datetime, timedelta
        global session_start, power
        if power and session_start and datetime.now() - session_start > timedelta(minutes=30):
            power = False
            var1.set('管理员未登录')
            messagebox.showwarning('会话过期', '管理员登录已过期，请重新登录')
            return
        if not power:
            tk.messagebox.showerror(title='Error', message='你还未登录,没有管理员权限')
        else:
            return func(*args, **kw)
    return wrapper

# 评估文本是否符合高级查询条件（支持正则、布尔、模糊）
def evaluate_query(text: str, query: str) -> bool:
    if not query:
        return True
    if query.startswith('/') and query.endswith('/'):
        try:
            return re.search(query[1:-1], text, re.IGNORECASE) is not None
        except re.error:
            return False
    q_upper = query.upper()
    if ' AND ' in q_upper:
        parts = re.split('(?i) AND ', query)
        return all(evaluate_query(text, p.strip()) for p in parts)
    if ' OR ' in q_upper:
        parts = re.split('(?i) OR ', query)
        return any(evaluate_query(text, p.strip()) for p in parts)
    if ' NOT ' in q_upper:
        parts = re.split('(?i) NOT ', query)
        include = parts[0].strip()
        exclude = parts[1].strip() if len(parts)>1 else ''
        return evaluate_query(text, include) and not evaluate_query(text, exclude)
    if query.lower() in text.lower():
        return True
    return SequenceMatcher(None, query.lower(), text.lower()).ratio() > 0.6

# 获取书籍摘要
def get_book_summary(book_name, author):
    try:
        api_url = "https://api.vveai.com/v1/chat/completions"
        api_key = "YOUR_API_KEY_HERE"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        data = {
            "model": "gpt-4o-mini",
            "messages": [
                {"role": "system", "content": "你是一位专业的图书馆员，请用简短的中文为读者提供书籍摘要信息。"},
                {"role": "user", "content": f"请为《{book_name}》(作者: {author})提供一个200字以内的中文摘要。"}
            ],
            "temperature": 0.7,
            "max_tokens": 300
        }
        resp = requests.post(api_url, headers=headers, data=json.dumps(data))
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"].strip()
        else:
            return f"获取摘要失败: HTTP {resp.status_code}"
    except Exception as e:
        return f"获取摘要时出错: {e}"

# 用户登录及注册
def usr():
    def usr_login():
        global var1, power, session_start
        uname = var_usr_name.get()
        pwd   = var_usr_pwd.get()
        try:
            with open('usrs_info.pickle','rb') as f:
                usrs = pickle.load(f)
            usrs['admin'] = 'adminadmin'
            with open('usrs_info.pickle','wb') as f:
                pickle.dump(usrs,f)
        except:
            usrs = {'admin':'adminadmin'}
            with open('usrs_info.pickle','wb') as f:
                pickle.dump(usrs,f)
        if uname in usrs and pwd==usrs[uname]:
            messagebox.showinfo('Welcome',f'管理员{uname}登录成功')
            power=True
            var1.set(f'管理员{uname}已登录')
            from datetime import datetime
            session_start = datetime.now()
            # 持久化会话状态
            with open(session_file, 'wb') as f:
                pickle.dump({'start': session_start.timestamp(), 'power': True}, f)
            log_action(f"管理员登录: {uname}")
            userwindow.destroy()
            build_sidebar()
            show_page('search')
        elif uname in usrs:
            messagebox.showerror('Error','密码错误')
        else:
            if messagebox.askyesno('Welcome','用户不存在，是否注册？'):
                usr_sign_up()
    @log
    def usr_sign_up():
        def do_sign():
            nn = new_name.get(); np=new_pwd.get(); npf=new_pwd_confirm.get()
            with open('usrs_info.pickle','rb') as f:
                exist=pickle.load(f)
            if np!=npf:
                messagebox.showerror('Error','两次密码不一致')
            elif nn in exist:
                messagebox.showerror('Error','用户已注册')
            else:
                exist[nn]=np
                with open('usrs_info.pickle','wb') as f:
                    pickle.dump(exist,f)
                messagebox.showinfo('Welcome','注册成功')
                signup_win.destroy()
        signup_win=tk.Toplevel(userwindow)
        signup_win.geometry('350x200'); signup_win.title('注册')
        tk.Label(signup_win,text='用户名:').place(x=10,y=10)
        new_name=tk.StringVar(); tk.Entry(signup_win,textvariable=new_name).place(x=150,y=10)
        tk.Label(signup_win,text='密码:').place(x=10,y=50)
        new_pwd=tk.StringVar(); tk.Entry(signup_win,textvariable=new_pwd,show='*').place(x=150,y=50)
        tk.Label(signup_win,text='确认密码:').place(x=10,y=90)
        new_pwd_confirm=tk.StringVar(); tk.Entry(signup_win,textvariable=new_pwd_confirm,show='*').place(x=150,y=90)
        tk.Button(signup_win,text='添加管理员',command=do_sign).place(x=150,y=130)
    userwindow=tk.Toplevel()
    userwindow.title('管理员登录'); userwindow.geometry('450x300+450+300'); userwindow.resizable(0,0)
    tk.Label(userwindow,text='请管理员登录\n大学图书馆\n图书馆管理系统',font=('Arial',20)).place(x=110,y=40)
    tk.Label(userwindow,text='用户名:').place(x=100,y=150)
    tk.Label(userwindow,text='密码:').place(x=100,y=190)
    var_usr_name=tk.StringVar(); tk.Entry(userwindow,textvariable=var_usr_name).place(x=160,y=150)
    var_usr_pwd =tk.StringVar(); tk.Entry(userwindow,textvariable=var_usr_pwd,show='*').place(x=160,y=190)
    tk.Button(userwindow,text='登录',command=usr_login).place(x=140,y=230)
    tk.Button(userwindow,text='注册',command=usr_sign_up).place(x=240,y=230)

def loginuser():
    usr()

# 导入图书
@log
def importbook_button():
    path=os.path.join(os.path.dirname(__file__),'book.xlsx')
    if not os.path.exists(path):
        messagebox.showerror('错误','未找到 book.xlsx')
        return
    wb=load_workbook(path); sh=wb.active
    dellist(tree)
    for r in sh.iter_rows(min_row=2,values_only=True):
        _,name,author,comp,id_,total,avail=r
        tree.insert('','end',
            values=[id_,name,author,comp,total,avail],
            tags=('highlight',))

# 启动时加载所有图书（绕过管理员鉴权）
def init_importbook():
    """加载所有图书列表，绕过管理员鉴权"""
    path = os.path.join(os.path.dirname(__file__), 'book.xlsx')
    if not os.path.exists(path):
        messagebox.showerror('错误', '未找到 book.xlsx')
        return
    wb = load_workbook(path); sh = wb.active
    dellist(tree)
    for r in sh.iter_rows(min_row=2, values_only=True):
        _, name, author, comp, id_, total, avail = r
        tree.insert('', 'end',
                    values=[id_, name, author, comp, total, avail],
                    tags=('highlight',))

# 批量导入按钮功能
@log
def batch_import_button():
    # 选择待导入的 Excel 文件
    file_path = filedialog.askopenfilename(
        title='选择批量导入的Excel文件',
        filetypes=[('Excel 文件', '*.xlsx')]
    )
    if not file_path:
        return
    try:
        # 读取目标表格
        wb_new = load_workbook(file_path)
        sh_new = wb_new.active
        # 读取现有图书表
        book_path = os.path.join(os.path.dirname(__file__), 'book.xlsx')
        wb = load_workbook(book_path)
        sh = wb.active
        # 构建现有书籍字典，key=id
        existing = { str(row[4].value): row for row in sh.iter_rows(min_row=2) }
        # 从新表导入数据，覆盖或新增
        for r in sh_new.iter_rows(min_row=2, values_only=True):
            _, name, author, comp, id_, total, avail = r
            if str(id_) in existing:
                # 覆盖该行
                row = existing[str(id_)]
                row[1].value = name
                row[2].value = author
                row[3].value = comp
                row[5].value = total
                row[6].value = avail
            else:
                # 新增行
                sh.append([None, name, author, comp, id_, total, avail])
        wb.save(book_path)
        messagebox.showinfo('成功', '批量导入完成')
        importbook_button()
    except Exception as e:
        messagebox.showerror('错误', f'批量导入失败: {e}')
# 查找图书
def search_button():
    q = search_name.get().strip()
    aq = search_author.get().strip()
    path=os.path.join(os.path.dirname(__file__),'book.xlsx')
    try:
        wb=load_workbook(path); sh=wb.active
        dellist(tree); found=False
        for r in sh.iter_rows(min_row=2,values_only=True):
            _,name,author,comp,id_,total,avail=r
            if evaluate_query(name,q) and evaluate_query(author,aq):
                tree.insert('','end',
                    values=[id_,name,author,comp,total,avail,'','',''],
                    tags=('highlight',))
                found=True
        if not found:
            messagebox.showinfo('提示','未找到匹配的图书')
    except Exception as e:
        messagebox.showerror('错误',f'搜索失败：{e}')

# 借出图书
def lendbook_button():
    sel=lb.get(0,'end')
    if not sel:
        messagebox.showwarning('警告','请先选中要借出的图书')
        return
    sid = stu_idEntry.get().strip()
    sname = stu_nameEntry.get().strip()
    # 校验学生学号和姓名是否在 name.xlsx 中
    name_path = os.path.join(os.path.dirname(__file__), 'name.xlsx')
    if not os.path.exists(name_path):
        messagebox.showerror('错误', '未找到名单文件')
        return
    wb_n = load_workbook(name_path)
    sh_n = wb_n.active
    valid = any(
        str(r[0]) == sid and str(r[2]) == sname
        for r in sh_n.iter_rows(min_row=2, values_only=True)
    )
    if not valid:
        messagebox.showerror('错误', '学号和姓名不匹配，请检查输入')
        return
    if not sid or not sname:
        # 弹出填写学生信息窗口
        def submit_info():
            stu_idEntry.set(entry_id.get().strip())
            stu_nameEntry.set(entry_name.get().strip())
            info_win.destroy()
            lendbook_button()
        info_win = tk.Toplevel(window1)
        info_win.title('填写学生信息')
        tk.Label(info_win, text='学生学号:').grid(row=0, column=0, padx=5, pady=5)
        entry_id = tk.StringVar()
        tk.Entry(info_win, textvariable=entry_id).grid(row=0, column=1, padx=5, pady=5)
        tk.Label(info_win, text='学生姓名:').grid(row=1, column=0, padx=5, pady=5)
        entry_name = tk.StringVar()
        tk.Entry(info_win, textvariable=entry_name).grid(row=1, column=1, padx=5, pady=5)
        tk.Button(info_win, text='提交', command=submit_info).grid(row=2, column=0, columnspan=2, pady=10)
        return
    path=os.path.join(os.path.dirname(__file__),'book.xlsx')
    wb=load_workbook(path); sh=wb.active; updated=False
    for itm in sel:
        bid,bname=itm.split('|',1) if '|' in itm else (None,None)
        for row in sh.iter_rows(min_row=2):
            if str(row[4].value)==bid:
                cell=row[6]
                if cell.value>0:
                    cell.value-=1; updated=True
                else:
                    messagebox.showwarning('警告',f'《{bname}》已无可借复本')
                break
    if updated:
        bp=os.path.join(os.path.dirname(__file__),'borrow.xlsx')
        if not os.path.exists(bp):
            wb_b=Workbook(); sh_b=wb_b.active
            sh_b.append(['学生学号','学生姓名','书籍编号','书籍名称','借书日期'])
        else:
            wb_b=load_workbook(bp); sh_b=wb_b.active
        for itm in sel:
            bid,bname=itm.split('|',1) if '|' in itm else (None,None)
            sh_b.append([sid,sname,bid,bname,time.strftime('%Y-%m-%d')])
        wb_b.save(bp)
        # 记录到持久化借阅历史
        import csv
        header = ['学生学号','学生姓名','书籍编号','书籍名称','借书日期']
        write_header = not os.path.exists(history_file)
        with open(history_file, 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if write_header:
                writer.writerow(header)
            for itm in sel:
                bid, bname = itm.split('|',1)
                writer.writerow([sid, sname, bid, bname, time.strftime('%Y-%m-%d')])
        wb.save(path)
        messagebox.showinfo('成功',f'学生<{sname}>借书成功')
        log_action(f"借出: 学生 {sname}({sid}) 借书 {bname}({bid})")
        dellb(); init_importbook()

# 归还图书
def returnbook_button():
    sid = stu_idEntry.get().strip()
    sel_idx = lb.curselection()              # 仅获取被选中的条目索引
    if not sel_idx:
        messagebox.showwarning('警告', '请先选中要归还的图书')
        return
    sel = [lb.get(i) for i in sel_idx]       # 列表推导获取选中的条目内容
    path = os.path.join(os.path.dirname(__file__), 'book.xlsx')
    wb = load_workbook(path); sh = wb.active; updated = False
    for itm in sel:
        bid, bname = itm.split('|', 1) if '|' in itm else (None, None)
        for row in sh.iter_rows(min_row=2):
            if str(row[4].value) == bid:
                row[6].value += 1; updated = True
                break
        bp = os.path.join(os.path.dirname(__file__), 'borrow.xlsx')
        if os.path.exists(bp):
            wb_b = load_workbook(bp); sh_b = wb_b.active
            for r in list(sh_b.iter_rows(min_row=2)):
                if str(r[0].value) == sid and str(r[2].value) == bid:
                    sh_b.delete_rows(r[0].row, 1)
            wb_b.save(bp)
    # 从列表框移除已归还的条目
    for i in reversed(sel_idx):
        lb.delete(i)
    if updated:
        wb.save(path)
        messagebox.showinfo('成功', '图书归还成功')
        log_action(f"归还: 学生 {stu_nameEntry.get().strip()} 归还书籍 {bid}")
        dellb(); init_importbook()

# 删除图书
@log
def removebook_button():
    if messagebox.askyesno('确认','是否删除选中图书？'):
        vals=lb.get(0,'end')
        for itm in vals:
            bid,_=itm.split('|',1)
            conn=connect(host='localhost',port=3306,user='root',password='YOUR_PASSWORD',database='library')
            cur=conn.cursor()
            cur.execute('DELETE FROM book WHERE book_id="%s";'%bid)
            conn.commit();cur.close();conn.close()
            log_action(f"删除: 管理员删除书籍 {bid}")
        lb.delete(0,'end')
        messagebox.showinfo('提示','删除成功')

# 编辑图书/读者
@log
def editbook_button():
    editbook()
@log
def editreader_button():
    editreader()

# 双击展示摘要和元数据
def treeviewClick(event):
    sel = tree.selection()
    if not sel:
        return
    vals = tree.item(sel[0], 'values')
    book_id, book_name, author = vals[0], vals[1], vals[2]
    summary_text.delete('1.0', tk.END)
    summary_text.insert(tk.END, f"正在获取《{book_name}》摘要...\n")
    def worker():
        s = get_book_summary(book_name, author)
        def ui():
            summary_text.delete('1.0', tk.END)
            summary_text.insert(tk.END, f"《{book_name}》摘要:\n\n{s}")
            # 生成仅包含书籍编号的二维码
            qr_img = qrcode.make(book_id)
            qr_img = qr_img.resize((150, 150))
            tk_qr = ImageTk.PhotoImage(qr_img)
            qr_label.configure(image=tk_qr)
            qr_label.image = tk_qr
        summary_text.after(0, ui)
    threading.Thread(target=worker, daemon=True).start()

# 添加到借阅列表功能
def add_to_borrow():
    sel = tree.selection()
    if not sel:
        messagebox.showwarning('提示', '请先选中要借阅的图书')
        return
    vals = tree.item(sel[0], 'values')
    book_id, book_name = vals[0], vals[1]
    lb.insert('end', f"{book_id}|{book_name}")

# 加载借阅列表
def load_borrowed():
    # 清空列表
    lb.delete(0, 'end')
    sid = stu_idEntry.get().strip()
    sname = stu_nameEntry.get().strip()
    if not sid or not sname:
        messagebox.showwarning('提示', '请先填写学生学号和姓名')
        return
    bp = os.path.join(os.path.dirname(__file__), 'borrow.xlsx')
    if not os.path.exists(bp):
        messagebox.showinfo('提示', '暂无借阅记录')
        return
    wb_b = load_workbook(bp); sh_b = wb_b.active
    for r in sh_b.iter_rows(min_row=2, values_only=True):
        if str(r[0]) == sid and str(r[1]) == sname:
            lb.insert('end', f"{r[2]}|{r[3]}")


# IC卡读取功能 (使用 Adafruit PN532_UART)
def read_ic_card():
    """使用 PN532 通过 UART 读取 IC 卡号并自动填充学号与姓名"""
    try:
        # 打开串口并初始化 PN532
        port = None
        for p in serial.tools.list_ports.comports():
            if 'usbmodem' in p.device or 'usbserial' in p.device:
                port = p.device
                break
        if not port:
            messagebox.showerror('错误', '未检测到 USB 串口设备')
            return
        uart = serial.Serial(port, baudrate=115200, timeout=1)
        pn532 = PN532_UART(uart, debug=False)
        pn532.SAM_configuration()
        # 读取卡片 UID（最大等待 5 秒）
        uid = pn532.read_passive_target(timeout=5)
        uart.close()
        if uid is None:
            messagebox.showwarning('提示', '未读取到卡号，请重试')
            return
        # 将 UID 转为十六进制字符串
        sid = ''.join('{:02X}'.format(i) for i in uid)
        # 在 name.xlsx 中查找对应姓名
        name_path = os.path.join(os.path.dirname(__file__), 'name.xlsx')
        if not os.path.exists(name_path):
            messagebox.showerror('错误', '未找到 name.xlsx 文件')
            return
        wb_n = load_workbook(name_path)
        sh_n = wb_n.active
        # name.xlsx 格式：学号(列1), 姓名(列2), 卡号(列3)
        for row in sh_n.iter_rows(min_row=2, values_only=True):
            # 根据卡号(第二列)匹配
            if str(row[1]) == sid:
                stu_idEntry.set(str(row[0]))
                stu_nameEntry.set(str(row[2]))
                return
        messagebox.showwarning('提示', f'UID {sid} 未在名单中找到')
    except Exception as e:
        messagebox.showerror('错误', f'读取卡失败: {e}')

# 清空
def dellb():
    lb.delete(0,'end')
def dellist(t):
    for x in t.get_children():
        t.delete(x)

# 查看学生借阅
def viewstudent():
    sid=stu_idEntry.get().strip(); sname=stu_nameEntry.get().strip()
    if not sid or not sname:
        messagebox.showwarning('警告','请输入学号和姓名'); return
    bp=os.path.join(os.path.dirname(__file__),'borrow.xlsx')
    if not os.path.exists(bp):
        messagebox.showinfo('提示','暂无借书记录'); return
    wb=load_workbook(bp); sh=wb.active
    w=tk.Toplevel(); w.title('学生借阅信息'); w.geometry('600x300+450+300')
    cols=['1','2','3','4','5']; texts=['学号','姓名','书籍编号','书名','借书日期']
    tv=ttk.Treeview(w,columns=cols,show='headings',height=14)
    for i,t in enumerate(texts,start=1):
        tv.heading(str(i),text=t); tv.column(str(i),width=100,anchor='center')
    tv.place(x=0,y=0)
    for r in sh.iter_rows(min_row=2,values_only=True):
        if str(r[0])==sid and str(r[1])==sname:
            tv.insert('','end',values=list(r))
    def on(evt):
        for it in tv.selection():
            v=tv.item(it,'values')
            lb.insert('end',f"{v[2]}|{v[3]}")
    tv.bind('<Double-1>',on)

# 逾期查看
@log
def overtime():
    bp=os.path.join(os.path.dirname(__file__),'borrow.xlsx')
    if not os.path.exists(bp):
        messagebox.showinfo('提示','暂无借书记录'); return
    wb=load_workbook(bp); sh=wb.active
    rec=[]; today=datetime.today().date()
    for r in sh.iter_rows(min_row=2,values_only=True):
        sid,sname,bid,bname,bdate=r
        try:
            bd=datetime.strptime(bdate,'%Y-%m-%d').date()
        except:
            continue
        due=bd+timedelta(days=30); days=(due-today).days
        rec.append((days,sid,sname,bid,bname,bd,due))
    rec.sort(key=lambda x:x[0])
    w=tk.Toplevel(); w.title('逾期名单'); w.geometry('700x350')
    cols=['天数','学号','姓名','书号','书名','借出','到期']
    tv=ttk.Treeview(w,columns=[str(i) for i in range(len(cols))],
                    show='headings',height=15)
    for i,c in enumerate(cols):
        tv.heading(str(i),text=c); tv.column(str(i),width=100,anchor='center')
    tv.place(x=0,y=0)
    for record in rec:
        tv.insert('', 'end', values=record)

# 日志查看
def book_log():
    lp = os.path.join(os.path.dirname(__file__), 'log.txt')
    if not os.path.exists(lp):
        messagebox.showinfo('日志', '无日志')
        return
    with open(lp, 'r', encoding='utf-8') as f:
        content = f.read()
    # 弹出日志查看窗口
    log_win = tk.Toplevel(window1)
    log_win.title('日志')
    log_win.geometry('600x400')
    st = scrolledtext.ScrolledText(log_win, width=80, height=20, wrap=tk.WORD)
    st.pack(fill='both', expand=True, padx=10, pady=10)
    st.insert(tk.END, content)
    st.config(state='disabled')

@log
def overuser():
    global var1, power
    power = False
    var1.set('管理员未登录')
    # 删除会话文件
    if os.path.exists(session_file):
        os.remove(session_file)
    log_action("管理员注销")
    build_sidebar()
    show_page('search')

def scan_return():
    # 必须先填写学生学号和姓名
    if not stu_idEntry.get().strip() or not stu_nameEntry.get().strip():
        messagebox.showwarning('警告', '请先填写学生学号和姓名')
        return

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror('错误', '无法打开摄像头')
        return

    # 创建扫描窗口
    scan_win = tk.Toplevel(window1)
    scan_win.title('扫码还书')
    scan_win.geometry('400x300')
    video_label = tk.Label(scan_win)
    video_label.pack()

    detector = cv2.QRCodeDetector()

    def video_loop():
        ret, frame = cap.read()
        if not ret:
            scan_win.after(10, video_loop)
            return
        # OpenCV BGR -> RGB
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(frame_rgb)
        img = img.resize((400, 300))
        tk_img = ImageTk.PhotoImage(img)
        video_label.config(image=tk_img)
        video_label.image = tk_img

        # 尝试解码二维码
        data, bbox, _ = detector.detectAndDecode(frame)
        if data:
            # 读取借还信息
            book_id = data.strip()
            stu_id = stu_idEntry.get().strip()
            stu_name = stu_nameEntry.get().strip()
            # 验证是否已借出此书
            borrow_path = os.path.join(os.path.dirname(__file__), 'borrow.xlsx')
            if os.path.exists(borrow_path):
                wb_b = load_workbook(borrow_path); sh_b = wb_b.active
                has_record = any(
                    str(r[0].value) == stu_id and str(r[2].value) == book_id
                    for r in sh_b.iter_rows(min_row=2)
                )
            else:
                has_record = False
            if not has_record:
                messagebox.showinfo('提示', '该学生未借出此书，无需扫码还书')
                cap.release()
                scan_win.destroy()
                return
            # 更新库存
            path = os.path.join(os.path.dirname(__file__), 'book.xlsx')
            wb = load_workbook(path); sh = wb.active
            for row in sh.iter_rows(min_row=2):
                if str(row[4].value) == book_id:
                    row[6].value += 1
                    # 确保可借复本不超过馆藏复本
                    if row[6].value > row[5].value:
                        row[6].value = row[5].value
                    break
            wb.save(path)
            # 删除借阅记录
            bp = os.path.join(os.path.dirname(__file__), 'borrow.xlsx')
            if os.path.exists(bp):
                wb_b = load_workbook(bp); sh_b = wb_b.active
                for r in list(sh_b.iter_rows(min_row=2)):
                    if str(r[0].value) == stu_id and str(r[2].value) == book_id:
                        sh_b.delete_rows(r[0].row, 1)
                wb_b.save(bp)
            messagebox.showinfo('成功', f'学生[{stu_name}]归还图书[{book_id}]成功')
            cap.release()
            scan_win.destroy()
            init_importbook()
            return
        scan_win.after(30, video_loop)

    video_loop()

# 借阅分析函数
def analyze_borrow_stats():
    # 读取借阅历史记录
    hp = history_file
    if not os.path.exists(hp):
        return None, None
    # 使用 CSV 读取历史记录
    df_hist = pd.read_csv(hp, encoding='utf-8')
    # 选择列名：尝试“书籍名称”，否则“书名”
    if '书籍名称' in df_hist.columns:
        name_col = '书籍名称'
    elif '书名' in df_hist.columns:
        name_col = '书名'
    else:
        return None, None
    # 按书名汇总借出次数
    data = df_hist[name_col].value_counts().to_dict()
    # 调用第三方 OpenAI 服务生成分析报告
    prompt = (f"请根据下列数据，给出图书馆借阅情况的分析报告，并提出改进建议：\n"
              f"{data}\n分析应在200字以内。")
    api_url = "https://api.vveai.com/v1/chat/completions"
    api_key = "YOUR_API_KEY_HERE"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    payload = {
        "model": "gpt-4o-mini",
        "messages": [
            {"role": "system", "content": "你是一位专业的图书馆管理员，请根据借阅数据生成分析报告，并提出改进建议。"},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.7,
        "max_tokens": 300
    }
    try:
        r = requests.post(api_url, headers=headers, json=payload, timeout=10)
        r.raise_for_status()
        report = r.json()["choices"][0]["message"]["content"].strip()
    except Exception as e:
        report = f"分析生成失败: {e}"
    # 将数据转为 DataFrame 并调用可视化 API 示例
    df = pd.DataFrame(list(data.items()), columns=['Book Name', 'Borrow Count'])
    # 使用 QuickChart 生成条形图
    config = {
        "type": "bar",
        "data": {
            "labels": df["Book Name"].tolist(),
            "datasets": [{
                "label": "Borrow Count",
                "data": df["Borrow Count"].tolist()
            }]
        }
    }
    qc_url = "https://quickchart.io/chart"
    try:
        qc_resp = requests.get(qc_url, params={"c": json.dumps(config)}, timeout=10)
        qc_resp.raise_for_status()
        img = Image.open(BytesIO(qc_resp.content))
    except Exception:
        img = None
    return report, img

# ---------- 界面初始化 (CustomTkinter) ----------
window1 = ctk.CTk()
# 管理员状态显示
var1 = tk.StringVar(value='未登录')
window1.title('图书管理系统')
window1.geometry('1200x700')
window1.resizable(False, False)

# 主布局：左侧导航栏，右侧内容区
sidebar = ctk.CTkFrame(window1, width=200, corner_radius=15)
sidebar.pack(side='left', fill='y', padx=10, pady=10)

content = ctk.CTkFrame(window1, corner_radius=15)
content.pack(side='right', fill='both', expand=True, padx=10, pady=10)

 # 定义不同页面的 Frame
pages = {}
for name in ['search', 'borrow', 'log', 'admin', 'analysis']:
    frame = ctk.CTkFrame(content, corner_radius=10)
    frame.place(relx=0, rely=0, relwidth=1, relheight=1)
    pages[name] = frame

def show_page(page_name):
    for f in pages.values():
        f.lower()
    pages[page_name].lift()

# ----------- 侧栏构建函数 -----------
def build_sidebar():
    # 清空现有侧栏
    for widget in sidebar.winfo_children():
        widget.destroy()
    # 导航标签
    ctk.CTkLabel(sidebar, text="导航", font=("", 16)).pack(pady=(10, 20))
    # 登录状态
    ctk.CTkLabel(sidebar, textvariable=var1, font=("", 14)).pack(pady=(0,10))
    # 普通功能
    for text, key in [('查询','search'),('借阅','borrow')]:
        ctk.CTkButton(sidebar, text=text,
                      command=lambda k=key: show_page(k),
                      corner_radius=10).pack(fill='x', padx=20, pady=5)
    # 管理员功能按钮
    if not power:
        ctk.CTkButton(sidebar, text='管理员登录',
                      command=loginuser,
                      corner_radius=10).pack(fill='x', padx=20, pady=5)
    else:
        ctk.CTkButton(sidebar, text='注销登录',
                      command=overuser,
                      corner_radius=10).pack(fill='x', padx=20, pady=5)
        for text, key in [('日志','log'),('管理','admin')]:
            ctk.CTkButton(sidebar, text=text,
                          command=lambda k=key: show_page(k),
                          corner_radius=10).pack(fill='x', padx=20, pady=5)
        ctk.CTkButton(sidebar, text='借阅分析',
                      command=lambda k='analysis': show_page(k),
                      corner_radius=10).pack(fill='x', padx=20, pady=5)
    # 退出
    ctk.CTkButton(sidebar, text='退出',
                  command=window1.destroy,
                  corner_radius=10, fg_color='tomato').pack(fill='x', padx=20, pady=5)

build_sidebar()

# 构建“查询”页面
frame = pages['search']
ctk.CTkLabel(frame, text="书籍检索", font=("", 18)).pack(pady=10)
search_name = ctk.CTkEntry(frame, placeholder_text="书名关键字")
search_name.pack(padx=20, pady=5, fill='x')
search_author = ctk.CTkEntry(frame, placeholder_text="作者关键字")
search_author.pack(padx=20, pady=5, fill='x')
# 搜索及添加按钮
search_btn_frame = ctk.CTkFrame(frame)
search_btn_frame.pack(pady=10)
ctk.CTkButton(search_btn_frame, text="搜索", command=search_button,
              corner_radius=10).pack(side='left', padx=5)
ctk.CTkButton(search_btn_frame, text="添加到借阅列表", command=add_to_borrow,
              corner_radius=10).pack(side='left', padx=5)
# 主表格
tree = ttk.Treeview(frame, columns=[str(i) for i in range(1,7)],
                    show='headings', height=15)
for i, (w, h) in enumerate([(125,'编号'),(175,'书名'),(100,'作者'),
                            (150,'出版社'),(75,'馆藏'),(75,'可借')], start=1):
    tree.column(str(i), width=w, anchor='center')
    tree.heading(str(i), text=h)
tree.tag_configure('highlight', background='#FFA500', foreground='black')
tree.bind('<Double-1>', treeviewClick)
tree.pack(side='left', padx=20, pady=10, fill='both', expand=True)
# 摘要 + 二维码
summary_frame = ctk.CTkFrame(frame, corner_radius=10)
summary_frame.pack(side='right', padx=20, pady=10, fill='y')
summary_text = ctk.CTkTextbox(summary_frame, height=12, wrap='word')
summary_text.pack(padx=5, pady=5, fill='both', expand=True)
qr_label = ctk.CTkLabel(summary_frame, text="")
qr_label.pack(pady=5)

# 构建“借阅”页面
frame = pages['borrow']
ctk.CTkLabel(frame, text="借阅操作", font=("", 18)).pack(pady=10)
# 借阅列表
lb = tk.Listbox(frame, width=40, height=8)
lb.pack(padx=20, pady=5)
stu_idEntry = tk.StringVar()
stu_nameEntry = tk.StringVar()
# 学生信息输入
stu_frame = ctk.CTkFrame(frame)
stu_frame.pack(padx=20, pady=5, fill='x')
ctk.CTkLabel(stu_frame, text="学号:", width=50, anchor='w').pack(side='left')
ctk.CTkEntry(stu_frame, textvariable=stu_idEntry, placeholder_text="请输入学号").pack(side='left', fill='x', expand=True, padx=(0,10))
ctk.CTkLabel(stu_frame, text="姓名:", width=50, anchor='w').pack(side='left')
ctk.CTkEntry(stu_frame, textvariable=stu_nameEntry, placeholder_text="请输入姓名").pack(side='left', fill='x', expand=True)
# 操作按钮一行
btn_frame_borrow = ctk.CTkFrame(frame)
btn_frame_borrow.pack(pady=10, padx=20, fill='x')
ctk.CTkButton(btn_frame_borrow, text="借出图书", command=lendbook_button, corner_radius=10).grid(row=0, column=0, padx=5)
ctk.CTkButton(btn_frame_borrow, text="清空列表", command=dellb, corner_radius=10).grid(row=0, column=1, padx=5)
ctk.CTkButton(btn_frame_borrow, text="加载借阅列表", command=load_borrowed, corner_radius=10).grid(row=0, column=2, padx=5)
ctk.CTkButton(btn_frame_borrow, text="还书", command=returnbook_button, corner_radius=10).grid(row=0, column=3, padx=5)
ctk.CTkButton(btn_frame_borrow, text="扫码还书", command=scan_return, corner_radius=10).grid(row=0, column=4, padx=5)
ctk.CTkButton(btn_frame_borrow, text="读取IC卡", command=read_ic_card, corner_radius=10).grid(row=0, column=5, padx=5)



# 构建“日志”页面
frame = pages['log']
ctk.CTkLabel(frame, text="操作日志", font=("", 18)).pack(pady=10)
log_text = ctk.CTkTextbox(frame, height=15, wrap='word')
log_text.pack(padx=20, pady=10, fill='both', expand=True)
def load_logs():
    lp = os.path.join(os.path.dirname(__file__), 'log.txt')
    if os.path.exists(lp):
        with open(lp, 'r', encoding='utf-8') as f:
            log_text.delete('0.0','end')
            log_text.insert('0.0', f.read())
ctk.CTkButton(frame, text="刷新日志", command=load_logs, corner_radius=10).pack(pady=5)

# 构建“管理”页面
frame = pages['admin']
ctk.CTkLabel(frame, text="管理操作", font=("", 18)).pack(pady=10)
ctk.CTkButton(frame, text="编辑图书", command=editbook_button, corner_radius=10).pack(pady=5)
ctk.CTkButton(frame, text="删除图书", command=removebook_button, corner_radius=10).pack(pady=5)
ctk.CTkButton(frame, text="批量导入", command=batch_import_button, corner_radius=10).pack(pady=5)

# 构建“模型分析”页面（仅管理员可见）
frame = pages['analysis']
ctk.CTkLabel(frame, text="借阅情况分析", font=("", 18)).pack(pady=10)
analysis_text = ctk.CTkTextbox(frame, height=15, wrap='word')
analysis_text.pack(padx=20, pady=(5,10), fill='both', expand=True)
analysis_img_label = ctk.CTkLabel(frame, text="")
analysis_img_label.pack(padx=20, pady=5)
def load_analysis():
    analysis_text.delete('0.0','end')
    analysis_text.insert('0.0','正在生成分析，请稍候...')
    def worker():
        report, img = analyze_borrow_stats()
        def ui():
            analysis_text.delete('0.0','end')
            analysis_text.insert('0.0', report or '无数据')
            if img:
                # 缩小可视化图表尺寸
                resized_img = img.resize((400, 300))
                tkimg = ImageTk.PhotoImage(resized_img)
                analysis_img_label.configure(image=tkimg)
                analysis_img_label.image = tkimg
        analysis_text.after(0, ui)
    threading.Thread(target=worker, daemon=True).start()
ctk.CTkButton(frame, text="刷新分析", command=load_analysis,
              corner_radius=10).pack(pady=10)


# 默认显示查询页面
show_page('search')
# 启动时列出所有图书（绕过权限检查）
init_importbook()

window1.mainloop()